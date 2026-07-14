import re
from openpyxl import load_workbook

def workbook(path): return load_workbook(path,data_only=True)

def product_info(ws):
 txt=str(ws["A1"].value)
 m=re.search(r"(.+?)\s+EP\s*Code\s*(\S+)",txt,re.I)
 if m:return m.group(1).replace("product","").strip(),m.group(2)
 return txt,""
