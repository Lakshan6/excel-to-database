"""
excel/reader.py
----------------
Opens the workbook and exposes its worksheets. We use openpyxl (not
pandas) so we can see the raw, un-filled merged-cell layout and forward
fill it ourselves with full control.
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook(path: str):
    """Load a workbook in read-only-safe mode (data_only=True resolves formulas)."""
    return openpyxl.load_workbook(path, data_only=True)


def iter_product_sheets(path: str):
    """Yield (sheet_name, worksheet) for every sheet in the workbook."""
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        yield sheet_name, wb[sheet_name]
