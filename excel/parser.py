"""
excel/parser.py
----------------
Turns a single worksheet (one product's test template) into structured
data: the product info from the title row, and a list of parameter rows
with merged-cell columns ("S.No." and "Test") forward-filled.
"""

import re
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet


def _normalize(text) -> str:
    """Collapse whitespace/newlines and lowercase, for robust header matching."""
    if text is None:
        return ""
    return " ".join(str(text).split()).strip().lower()


@dataclass
class ProductInfo:
    prod_name: str
    ep_code: Optional[str]
    prod_full_name: Optional[str] = None


@dataclass
class ParameterRow:
    s_no: Optional[str]
    test: Optional[str]
    parameter: str
    observation: Optional[str]
    remarks: Optional[str]


def parse_title(ws: Worksheet, title_pattern: str) -> ProductInfo:
    """
    Row 1 holds a single title such as 'product abc EP Code 123456789'.
    Find the first non-empty cell in row 1 and split it into
    prod_name / ep_code using the configured regex.
    """
    title_text = None
    for cell in ws[1]:
        if cell.value not in (None, ""):
            title_text = str(cell.value).strip()
            break

    if not title_text:
        raise ValueError(f"Sheet '{ws.title}': row 1 title is empty, cannot parse product info.")

    match = re.match(title_pattern, title_text, flags=re.IGNORECASE)
    if not match:
        # Fall back: use the whole title as prod_name, no ep_code found.
        return ProductInfo(prod_name=title_text, ep_code=None)

    groups = match.groupdict()
    return ProductInfo(
        prod_name=groups.get("prod_name", title_text).strip(),
        ep_code=(groups.get("ep_code") or "").strip() or None,
    )


def _locate_header_columns(ws: Worksheet, header_row: int, columns_map: Dict[str, str]) -> Dict[str, int]:
    """
    Scan the header row and match each configured logical column name
    (e.g. 'parameter') to its actual column index in the sheet.
    """
    normalized_targets = {key: _normalize(label) for key, label in columns_map.items()}

    found: Dict[str, int] = {}
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        normalized_cell = _normalize(cell_value)
        if not normalized_cell:
            continue
        for key, target in normalized_targets.items():
            if key in found:
                continue
            if normalized_cell == target:
                found[key] = col_idx

    missing = set(normalized_targets) - set(found)
    if missing:
        raise ValueError(
            f"Sheet '{ws.title}': could not locate header column(s) {missing} "
            f"in row {header_row}. Check mappings/column_mapping.json."
        )
    return found


def parse_parameter_rows(
    ws: Worksheet,
    header_row: int,
    data_start_row: int,
    columns_map: Dict[str, str],
) -> List[ParameterRow]:
    """
    Read every data row from data_start_row to the last populated row,
    forward-filling the merged 'S.No.' and 'Test' columns, and returning
    one ParameterRow per row that has a non-empty Parameter cell.
    """
    col_idx = _locate_header_columns(ws, header_row, columns_map)

    rows: List[ParameterRow] = []
    last_s_no = None
    last_test = None

    for row_idx in range(data_start_row, ws.max_row + 1):
        s_no_val = ws.cell(row=row_idx, column=col_idx["s_no"]).value
        test_val = ws.cell(row=row_idx, column=col_idx["test"]).value
        parameter_val = ws.cell(row=row_idx, column=col_idx["parameter"]).value
        observation_val = ws.cell(row=row_idx, column=col_idx["observation"]).value
        remarks_val = ws.cell(row=row_idx, column=col_idx["remarks"]).value

        # Forward fill merged cells
        if s_no_val not in (None, ""):
            last_s_no = str(s_no_val).strip()
        if test_val not in (None, ""):
            last_test = str(test_val).strip()

        if parameter_val in (None, ""):
            # Skip fully blank rows (e.g. trailing empty rows at sheet end)
            continue

        rows.append(
            ParameterRow(
                s_no=last_s_no,
                test=last_test,
                parameter=str(parameter_val).strip(),
                observation=str(observation_val).strip() if observation_val not in (None, "") else None,
                remarks=str(remarks_val).strip() if remarks_val not in (None, "") else None,
            )
        )

    return rows
